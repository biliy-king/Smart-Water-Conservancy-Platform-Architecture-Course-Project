import os
import sys
import openpyxl
from datetime import datetime
from django.utils import timezone

# 1. 配置Django环境（必须，否则无法使用ORM）
# 替换为你的Django项目根目录（比如你的项目叫smart_water）
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "hydro_platform.settings")  # 替换为你的settings模块路径
import django
django.setup()


# 2. 导入你的Django Model（替换为你实际的Model路径）
from water_structures.models import Structure, MonitoringDevice, Point
from monitoring.models import MonitorData



def step1_import_structure(excel_path):
    """
    第1步：导入大坝基础信息
    只需执行一次，创建唯一的大坝记录
    """
    print("\n===== 步骤1：导入大坝基础信息 =====")
    
    # 使用 get_or_create 确保不会重复创建
    dam, created = Structure.objects.get_or_create(
        name="河海大坝",  # 修改为你的大坝名称
        defaults={
            "cesium_center_x": 1000.0,  # 修改为实际坐标
            "cesium_center_y": 500.0,
            "cesium_center_z": 100.0,
            "cesium_heading": 0.0,
            "cesium_pitch": 0.0,
            "cesium_roll": 0.0,
            "cesium_scale": 1.0,
            "level": "2级",
            "completion_time": datetime(2026, 1, 15)  # 如果知道建成时间，填写如 "2010-06-15"
        }
    )
    
    if created:
        print(f"✅ 大坝创建成功：{dam.name} (ID: {dam.id})")
    else:
        print(f"⚠️  大坝已存在：{dam.name} (ID: {dam.id})，跳过创建")
    
    return dam


def step2_import_devices(excel_path):
    """
    第2步：导入监测设备
    根据Excel中的设备信息创建MonitoringDevice记录
    """
    print("\n===== 步骤2：导入监测设备 =====")
    
    # 获取大坝
    try:
        dam = Structure.objects.first()
        if not dam:
            print("❌ 错误：请先执行步骤1导入大坝")
            return
    except Exception as e:
        print(f"❌ 获取大坝失败：{e}")
        return
    
    # 定义设备清单（根据你的Excel和实际情况调整）
    devices_data = [
    # 倒垂线设备
    {
        "name": "IP1", 
        "type": "inverted_plumb_left_right", 
        "position": "1号坝段",
        "install_time": datetime(2011, 12, 17),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP1", 
        "type": "inverted_plumb_up_down", 
        "position": "1号坝段",
        "install_time": datetime(2011, 12, 17),
        "device_status": "running"
    },
    {
        "name": "IP3", 
        "type": "inverted_plumb_left_right", 
        "position": "2号坝段",
        "install_time": datetime(2011, 3, 26),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP3", 
        "type": "inverted_plumb_up_down", 
        "position": "2号坝段",
        "install_time": datetime(2011, 3, 26),
        "device_status": "running"
    },
    {
        "name": "IP5", 
        "type": "inverted_plumb_left_right", 
        "position": "3号坝段",
        "install_time": datetime(2011, 3, 26),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP5", 
        "type": "inverted_plumb_up_down", 
        "position": "3号坝段",
        "install_time": datetime(2011, 3, 26),
        "device_status": "running"
    },
    {
        "name": "IP9", 
        "type": "inverted_plumb_left_right", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 13),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP9", 
        "type": "inverted_plumb_up_down", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 13),
        "device_status": "running"
    },
    {
        "name": "IP7", 
        "type": "inverted_plumb_left_right", 
        "position": "5号坝段",
        "install_time": datetime(2014, 7, 14),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP7", 
        "type": "inverted_plumb_up_down", 
        "position": "5号坝段",
        "install_time": datetime(2014, 7, 14),
        "device_status": "running"
    },
    {
        "name": "IP8", 
        "type": "inverted_plumb_left_right", 
        "position": "6号坝段",
        "install_time": datetime(2014, 7, 24),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP8", 
        "type": "inverted_plumb_up_down", 
        "position": "6号坝段",
        "install_time": datetime(2014, 7, 24),
        "device_status": "running"
    },
    {
        "name": "IP6", 
        "type": "inverted_plumb_left_right", 
        "position": "7号坝段",
        "install_time": datetime(2011, 3, 26),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP6", 
        "type": "inverted_plumb_up_down", 
        "position": "7号坝段",
        "install_time": datetime(2011, 3, 26),
        "device_status": "running"
    },
    {
        "name": "IP4", 
        "type": "inverted_plumb_left_right", 
        "position": "9号坝段",
        "install_time": datetime(2011, 3, 26),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP4", 
        "type": "inverted_plumb_up_down", 
        "position": "9号坝段",
        "install_time": datetime(2011, 3, 26),
        "device_status": "running"
    },
    {
        "name": "IP2", 
        "type": "inverted_plumb_left_right", 
        "position": "10号坝段",
        "install_time": datetime(2011, 12, 17),  # 加上安装时间
        "device_status": "running"  # 运行状态：running/stopped/faulty
    },
    {
        "name": "IP2", 
        "type": "inverted_plumb_up_down", 
        "position": "10号坝段",
        "install_time": datetime(2011, 12, 17),
        "device_status": "running"
    },
    # 引张线设备
    {
        "name": "EX1-2", 
        "type": "tension_wire_up_down", 
        "position": "2号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-3", 
        "type": "tension_wire_up_down", 
        "position": "3号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-4", 
        "type": "tension_wire_up_down", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-5", 
        "type": "tension_wire_up_down", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-6", 
        "type": "tension_wire_up_down", 
        "position": "5号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-7", 
        "type": "tension_wire_up_down", 
        "position": "5号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-8", 
        "type": "tension_wire_up_down", 
        "position": "6号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-9", 
        "type": "tension_wire_up_down", 
        "position": "7号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-10", 
        "type": "tension_wire_up_down", 
        "position": "8号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX1-11", 
        "type": "tension_wire_up_down", 
        "position": "9号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX2-2", 
        "type": "tension_wire_up_down", 
        "position": "3号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX2-3", 
        "type": "tension_wire_up_down", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX2-4", 
        "type": "tension_wire_up_down", 
        "position": "5号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX2-5", 
        "type": "tension_wire_up_down", 
        "position": "6号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX2-6", 
        "type": "tension_wire_up_down", 
        "position": "7号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX2-7", 
        "type": "tension_wire_up_down", 
        "position": "8号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX3-2", 
        "type": "tension_wire_up_down", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX3-3", 
        "type": "tension_wire_up_down", 
        "position": "5号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX3-4", 
        "type": "tension_wire_up_down", 
        "position": "6号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    {
        "name": "EX3-4\'", 
        "type": "tension_wire_up_down", 
        "position": "7号坝段",
        "install_time": datetime(2018, 5, 10),
        "device_status": "running"
    },
    # 静力水准设备
    {
        "name": "TC1-1", 
        "type": "hydrostatic_leveling", 
        "position": "1号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-2", 
        "type": "hydrostatic_leveling", 
        "position": "2号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-3", 
        "type": "hydrostatic_leveling", 
        "position": "3号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-4", 
        "type": "hydrostatic_leveling", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-5", 
        "type": "hydrostatic_leveling", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-6", 
        "type": "hydrostatic_leveling", 
        "position": "5号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-6\'", 
        "type": "hydrostatic_leveling", 
        "position": "5号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-7", 
        "type": "hydrostatic_leveling", 
        "position": "5号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-8", 
        "type": "hydrostatic_leveling", 
        "position": "6号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-9", 
        "type": "hydrostatic_leveling", 
        "position": "7号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-10", 
        "type": "hydrostatic_leveling", 
        "position": "8号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-11", 
        "type": "hydrostatic_leveling", 
        "position": "9号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC1-12", 
        "type": "hydrostatic_leveling", 
        "position": "10号坝段",
        "install_time": datetime(2018, 5, 2),
        "device_status": "running"
    },
    {
        "name": "TC3-1", 
        "type": "hydrostatic_leveling", 
        "position": "3号坝段",
        "install_time": datetime(2018, 5, 5),
        "device_status": "running"
    },
    {
        "name": "TC3-2", 
        "type": "hydrostatic_leveling", 
        "position": "4号坝段",
        "install_time": datetime(2018, 5, 5),
        "device_status": "running"
    },
    {
        "name": "TC3-3", 
        "type": "hydrostatic_leveling", 
        "position": "5号坝段",
        "install_time": datetime(2018, 5, 5),
        "device_status": "running"
    },
    {
        "name": "TC3-4", 
        "type": "hydrostatic_leveling", 
        "position": "6号坝段",
        "install_time": datetime(2018, 5, 5),
        "device_status": "running"
    },
    {
        "name": "TC3-5", 
        "type": "hydrostatic_leveling", 
        "position": "7号坝段",
        "install_time": datetime(2018, 5, 5),
        "device_status": "running"
    },
    # 水位传感器设备
    {
        "name": "上游", 
        "type": "water_level_upstream", 
        "position": "上游",
        "install_time": None,
        "device_status": "running"
    },
    {
        "name": "下游", 
        "type": "water_level_downstream", 
        "position": "下游",
        "install_time": None,
        "device_status": "running"
    },
]
    
    success_count = 0
    for device_info in devices_data:
        try:
            device, created = MonitoringDevice.objects.get_or_create(
                structure=dam,
                device_name=device_info["name"],
                device_type=device_info["type"],
                defaults={
                    # 为避免唯一约束(structure, install_position, device_type)冲突，
                    # 将安装位置拓展为“坝段-设备名”，保证不同设备的安装位置唯一
                    "install_position": f"{device_info['position']}-{device_info['name']}",
                    "install_time": device_info.get("install_time"),      # 加这行
                    "device_status": device_info.get("device_status", "running"),  # 加这行
                }
            )
            if created:
                print(f"✅ 创建设备：{device.device_name}")
                success_count += 1
            else:
                print(f"⚠️  设备已存在：{device.device_name}")
        except Exception as e:
            print(f"❌ 创建设备失败：{device_info['name']} - {e}")
    
    print(f"\n设备导入完成：成功创建 {success_count} 个设备")


def step3_import_points(excel_path):
    """
    第3步：导入测点
    为每个设备创建对应的测点，设置相对坐标和阈值
    """
    print("\n===== 步骤3：导入测点 =====")
    
    # 动态生成测点清单模板：根据设备类型自动生成point_code
    # - 倒垂线设备会自动加上方向后缀（左右/上下），对应Excel列标题
    # - 其他设备point_code就等于device_name
    points_data = []
    for device in MonitoringDevice.objects.all():
        # 根据设备类型生成point_code
        if "inverted_plumb_left_right" in device.device_type:
            point_code = f"{device.device_name}-左右岸CH1"
        elif "inverted_plumb_up_down" in device.device_type:
            point_code = f"{device.device_name}-上下游CH2"
        elif device.device_type == "hydrostatic_leveling":
            # 静力水准表格列标题为“位移mm”，组合设备名+列标题以匹配Excel
            point_code = f"{device.device_name}-位移mm"
        elif device.device_type == "tension_wire_up_down":
            # 引张线表格同样使用“位移mm”列标题
            point_code = f"{device.device_name}-位移mm"
        else:
            # 其他设备直接用device_name作为point_code
            point_code = device.device_name
        
        points_data.append({
            "point_code": point_code,  # 与Excel列标题一致
            "device_name": device.device_name,
            "device_type": device.device_type,
            "relative_x": 0.0,
            "relative_y": 0.0,
            "relative_z": 0.0,
            "displacement_upper": None,
            "displacement_lower": None,
            "settlement_upper": None,
            "settlement_lower": None,
            "water_level_upper": None,
            "water_level_lower": None,
        })
    if not points_data:
        print("❌ 未找到设备，请先执行步骤2导入设备")
        return
    
    success_count = 0
    for point_info in points_data:
        try:
            # 查找对应的设备（按 device_name + device_type 唯一）
            device = MonitoringDevice.objects.get(device_name=point_info["device_name"], device_type=point_info["device_type"])
            desired_code = point_info["point_code"]

            # 如果该设备已绑定测点，则更新其 point_code 以匹配Excel标题
            existing_point = Point.objects.filter(device=device).first()
            if existing_point:
                if existing_point.point_code != desired_code:
                    existing_point.point_code = desired_code
                    existing_point.save()
                    print(f"🔄 更新测点编号：{desired_code}")
                else:
                    print(f"⚠️  测点已存在：{existing_point.point_code}")
                continue

            # 未绑定测点则创建
            point = Point.objects.create(
                device=device,
                point_code=desired_code,
                relative_x=point_info["relative_x"],
                relative_y=point_info["relative_y"],
                relative_z=point_info["relative_z"],
                displacement_upper=point_info.get("displacement_upper"),
                displacement_lower=point_info.get("displacement_lower"),
                settlement_upper=point_info.get("settlement_upper"),
                settlement_lower=point_info.get("settlement_lower"),
                water_level_upper=point_info.get("water_level_upper"),
                water_level_lower=point_info.get("water_level_lower"),
            )
            print(f"✅ 创建测点：{point.point_code}")
            success_count += 1
        except MonitoringDevice.DoesNotExist:
            print(f"❌ 设备不存在：{point_info['device_name']}，请先执行步骤2")
        except Exception as e:
            print(f"❌ 创建/更新测点失败：{point_info['point_code']} - {e}")
    
    print(f"\n测点导入完成：成功创建 {success_count} 个测点")


def step4_import_monitor_data(excel_path, sheet_name, device_type_field, time_col=1, header_row=7, data_start_row=8, data_col_start=4, device_name_row=None):
    """
    第4步:导入监测数据
    从Excel的指定Sheet读取数据，写入MonitorData表
    
    参数：
    - excel_path: Excel文件路径
    - sheet_name: Sheet名称（如"倒垂线"、"静力水准"）
    - device_type_field: 监测数据字段名（如"inverted_plumb_up_down"）
    - time_col: 监测时间所在列号（默认1=A列）
    - header_row: 列标题行号（默认7）
    - data_start_row: 数据起始行号（默认8）
    - data_col_start: 数据列起始位置（默认4=D列）
    - device_name_row: 设备名行号（倒垂线专用，默认None。如果设置，会组合设备名+列标题）
    """
    print(f"\n===== 步骤4：导入监测数据 - {sheet_name} ({device_type_field}) =====")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    
    # 动态构建测点映射（根据Excel列标题）
    point_mapping = {}
    for col_idx in range(data_col_start, ws.max_column + 1):  # 从指定列开始
        # 读取列标题
        column_header = ws.cell(row=header_row, column=col_idx).value
        column_header_str = str(column_header).strip() if column_header is not None else ""
        
        # 规范化全角单引号为半角（Excel 可能含有 U+2032 全角单引号或 U+FF07 全角撇号）
        column_header_str = column_header_str.replace('′', "'").replace('＇', "'")

        # 依据 sheet 与字段类型筛选需要的列，避免将不同类型写入同一字段
        # 倒垂线：CH1 → 左右岸；CH2 → 上下游
        if sheet_name == "倒垂线":
            if device_type_field == "inverted_plumb_left_right" and "左右岸CH1" not in column_header_str:
                continue
            if device_type_field == "inverted_plumb_up_down" and "上下游CH2" not in column_header_str:
                continue
        # 水位：上游/下游两列分别导入对应字段
        if sheet_name == "水位":
            if device_type_field == "water_level_upstream" and "上游" not in column_header_str:
                continue
            if device_type_field == "water_level_downstream" and "下游" not in column_header_str:
                continue
        # 静力水准/引张线：只取“位移mm”列
        if sheet_name in ("静力水准", "引张线"):
            if "位移mm" not in column_header_str:
                continue
        
        if column_header:
            # 如果有设备名行（如倒垂线），组合设备名+方向作为 point_code
            if device_name_row:
                device_name = ws.cell(row=device_name_row, column=col_idx).value
                if device_name and str(device_name).strip():
                    device_name_str = str(device_name).strip().replace('′', "'").replace('＇', "'")
                    point_code = f"{device_name_str}-{column_header_str}"
                else:
                    # 如果设备名为空，说明是合并单元格，向左查找
                    for left_col in range(col_idx - 1, data_col_start - 1, -1):
                        device_name = ws.cell(row=device_name_row, column=left_col).value
                        if device_name and str(device_name).strip():
                            device_name_str = str(device_name).strip().replace('′', "'").replace('＇', "'")
                            point_code = f"{device_name_str}-{column_header_str}"
                            break
                    else:
                        point_code = column_header_str
            else:
                # 其他 sheet 直接用列标题
                point_code = column_header_str
            
            try:
                point = Point.objects.get(point_code=point_code)
                point_mapping[col_idx] = point
                print(f"  映射列{col_idx} → 测点 {point.point_code}")
            except Point.DoesNotExist:
                print(f"  ⚠️  跳过列{col_idx}：测点 '{point_code}' 不存在")
    
    if not point_mapping:
        print("❌ 未找到任何有效测点映射，请检查Excel列标题是否与数据库一致")
        return
    
    # 遍历数据行
    success_count = 0
    fail_count = 0
    skip_count = 0
    invalid_data_count = 0  # 无效数据统计
    early_time_count = 0  # 早于安装时间的记录数
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
        # 解析监测时间
        monitor_time_str = row[time_col - 1]
        if not monitor_time_str:
            continue
        
        try:
            # 尝试多种时间格式
            if isinstance(monitor_time_str, datetime):
                monitor_time = monitor_time_str
            else:
                for fmt in ["%Y/%m/%d", "%Y-%m-%d", "%Y/%m/%d %H:%M:%S"]:
                    try:
                        monitor_time = datetime.strptime(str(monitor_time_str), fmt)
                        break
                    except:
                        continue
            monitor_time = timezone.make_aware(monitor_time)
        except Exception as e:
            fail_count += 1
            print(f"  ⚠️  行{row_idx}：时间格式错误 - {monitor_time_str}")
            continue
        
        # 遍历数据列
        for col_idx, point in point_mapping.items():
            value = row[col_idx - 1]
            
            # 1. 检查空值和 Excel 错误引用
            if value is None or value == "":
                skip_count += 1
                continue
            
            value_str = str(value).strip()
            
            # 2. 特殊处理：#REF! 当空白跳过
            if value_str == "#REF!":
                skip_count += 1
                continue
            
            # 3. 检查监测时间是否早于设备安装时间（重要：安装时间和记录时间必须对齐）
            # 设备安装时间为 DateField，监测时间为 datetime；使用日期对比避免类型错误
            if point.device.install_time and monitor_time.date() < point.device.install_time:
                early_time_count += 1
                continue
            
            # 4. 尝试将值转换为数字，特殊处理水位相关异常
            # 特殊值约定：
            # -999.1: 低于标尺水位/无法读数
            # -999.2: 被遮挡/无法观测
            # -999.9: 其他乱码/无法解析
            try:
                # 处理文字说明：根据关键词分类
                if "无法读数" in value_str or "低于标尺" in value_str or "低于水位" in value_str:
                    numeric_value = -999.1
                    invalid_data_count += 1
                    print(f"  ⚠️  行{row_idx} 列{col_idx}：低于标尺，用 -999.1 记录 - {value_str}")
                elif "无法观测" in value_str or "遮挡" in value_str or "杂物" in value_str:
                    numeric_value = -999.2
                    invalid_data_count += 1
                    print(f"  ⚠️  行{row_idx} 列{col_idx}：被遮挡，用 -999.2 记录 - {value_str}")
                # 处理 < 和 > 符号（水位特殊处理：直接取数字）
                elif value_str.startswith("<") or value_str.startswith(">") or value_str.startswith("＜") or value_str.startswith("＞") or value_str.startswith("≈"):
                    # 提取数字部分
                    numeric_part = ''.join(c for c in value_str if c.isdigit() or c == '.' or c == '-')
                    if numeric_part:
                        numeric_value = float(numeric_part)
                        print(f"  ℹ️  行{row_idx} 列{col_idx}：提取边界值 {value_str} → {numeric_value}")
                    else:
                        # 无法提取数字，用特殊值 -999.9 记录
                        numeric_value = -999.9
                        invalid_data_count += 1
                        print(f"  ⚠️  行{row_idx} 列{col_idx}：无法解析，用 -999.9 记录 - {value_str}")
                else:
                    # 正常数字
                    numeric_value = float(value_str)
            except (ValueError, TypeError):
                # 无法转换为数字（其他乱码），用特殊值 -999.9 记录
                numeric_value = -999.9
                invalid_data_count += 1
                print(f"  ⚠️  行{row_idx} 列{col_idx}：乱码数据，用 -999.9 记录 - {value}")
            
            # 5. 创建监测数据
            try:
                # 记录是否已存在（point + monitor_time唯一）
                existing = MonitorData.objects.filter(point=point, monitor_time=monitor_time).first()
                if existing:
                    # 若目标字段为空，则更新该字段；否则视为重复，跳过
                    current_val = getattr(existing, device_type_field)
                    if current_val is None:
                        setattr(existing, device_type_field, numeric_value)
                        existing.monitor_person = "Excel导入"
                        existing.save()
                        success_count += 1
                    else:
                        skip_count += 1
                        continue
                else:
                    data_dict = {
                        "point": point,
                        "monitor_time": monitor_time,
                        device_type_field: numeric_value,
                        "monitor_person": "Excel导入"
                    }
                    MonitorData.objects.create(**data_dict)
                    success_count += 1
            except Exception as e:
                fail_count += 1
                print(f"  ❌ 行{row_idx} 列{col_idx}：数据库错误 - {e}")
    
    print(f"\n{sheet_name} 导入完成：")
    print(f"  ✅ 成功：{success_count} 条")
    print(f"  ⚠️  跳过（空值）：{skip_count} 条")
    print(f"  ⚠️  无效数据（无法转换）：{invalid_data_count} 条")
    print(f"  ⚠️  早于安装时间：{early_time_count} 条")
    print(f"  ❌ 失败：{fail_count} 条")

def step5_view_data():
    """
    第5步：查看导入的数据
    显示各表的数据统计和样例
    """
    print("\n===== 数据查看 =====\n")
    
    # 1. 查看大坝
    print("【1. 大坝信息】")
    dams = Structure.objects.all()
    if dams.exists():
        for dam in dams:
            print(f"  ✓ {dam.name} (ID: {dam.id}, 等级: {dam.level})")
    else:
        print("  ✗ 未找到大坝记录")
    
    # 2. 查看设备
    print("\n【2. 监测设备】")
    devices = MonitoringDevice.objects.all()
    if devices.exists():
        device_types = devices.values('device_type').annotate(count=models.Count('id'))
        for dt in device_types:
            print(f"  ✓ {dt['device_type']}: {dt['count']} 个")
        print(f"  总计: {devices.count()} 个设备")
        print(f"\n  样例设备：")
        for device in devices[:5]:
            print(f"    - {device.device_name} ({device.device_type}, 位置: {device.install_position})")
    else:
        print("  ✗ 未找到设备记录")
    
    # 3. 查看测点
    print("\n【3. 测点】")
    points = Point.objects.all()
    if points.exists():
        print(f"  ✓ 总计: {points.count()} 个测点")
        print(f"\n  样例测点：")
        for point in points[:10]:
            print(f"    - {point.point_code} → 设备: {point.device.device_name}")
    else:
        print("  ✗ 未找到测点记录")
    
    # 4. 查看监测数据
    print("\n【4. 监测数据】")
    monitor_data = MonitorData.objects.all()
    if monitor_data.exists():
        print(f"  ✓ 总计: {monitor_data.count()} 条监测记录")
        
        # 统计各字段的数据量
        print("\n  各类型数据统计：")
        if monitor_data.filter(inverted_plumb_left_right__isnull=False).exists():
            count = monitor_data.filter(inverted_plumb_left_right__isnull=False).count()
            print(f"    - 倒垂线左右: {count} 条")
        if monitor_data.filter(inverted_plumb_up_down__isnull=False).exists():
            count = monitor_data.filter(inverted_plumb_up_down__isnull=False).count()
            print(f"    - 倒垂线上下: {count} 条")
        if monitor_data.filter(hydrostatic_leveling_settlement__isnull=False).exists():
            count = monitor_data.filter(hydrostatic_leveling_settlement__isnull=False).count()
            print(f"    - 静力水准: {count} 条")
        if monitor_data.filter(tension_wire_up_down__isnull=False).exists():
            count = monitor_data.filter(tension_wire_up_down__isnull=False).count()
            print(f"    - 张引线: {count} 条")
        if monitor_data.filter(water_level_upstream__isnull=False).exists():
            count = monitor_data.filter(water_level_upstream__isnull=False).count()
            print(f"    - 上游水位: {count} 条")
        if monitor_data.filter(water_level_downstream__isnull=False).exists():
            count = monitor_data.filter(water_level_downstream__isnull=False).count()
            print(f"    - 下游水位: {count} 条")
        
        # 检查异常数据
        print("\n  异常数据统计：")
        invalid_count = monitor_data.filter(
            models.Q(inverted_plumb_left_right__lt=-999) |
            models.Q(inverted_plumb_up_down__lt=-999) |
            models.Q(hydrostatic_leveling_settlement__lt=-999) |
            models.Q(tension_wire_up_down__lt=-999) |
            models.Q(water_level_upstream__lt=-999) |
            models.Q(water_level_downstream__lt=-999)
        ).count()
        if invalid_count > 0:
            print(f"    ⚠️  包含 -999.x 标记的异常数据: {invalid_count} 条")
            print(f"       (需要手动修正)")
        else:
            print(f"    ✓ 无异常数据标记")
        
        # 显示样例数据
        print(f"\n  最新5条监测数据：")
        for data in monitor_data.order_by('-monitor_time')[:5]:
            fields = []
            if data.inverted_plumb_left_right is not None:
                fields.append(f"左右={data.inverted_plumb_left_right}")
            if data.inverted_plumb_up_down is not None:
                fields.append(f"上下={data.inverted_plumb_up_down}")
            if data.hydrostatic_leveling_settlement is not None:
                fields.append(f"水准={data.hydrostatic_leveling_settlement}")
            if data.tension_wire_up_down is not None:
                fields.append(f"引张={data.tension_wire_up_down}")
            if data.water_level_upstream is not None:
                fields.append(f"上游水位={data.water_level_upstream}")
            if data.water_level_downstream is not None:
                fields.append(f"下游水位={data.water_level_downstream}")
            
            print(f"    {data.monitor_time.strftime('%Y-%m-%d %H:%M')} | {data.point.point_code} | {', '.join(fields)}")
    else:
        print("  ✗ 未找到监测数据")
    
    print("\n" + "=" * 80)


def step5_view_data():
    """
    步骤5：查看导入的数据统计
    用于验证导入结果是否正确
    """
    from django.db.models import Q
    
    print("\n" + "="*60)
    print("📊 查看导入数据统计")
    print("="*60 + "\n")
    
    # 1. 查看建筑物信息
    print("【1. 建筑物信息】")
    structures = Structure.objects.all()
    print(f"  建筑物数量: {structures.count()}")
    for s in structures:
        print(f"  - {s.name} (ID: {s.id})")
        print(f"    位置: cesium_center_x={s.cesium_center_x}, cesium_center_y={s.cesium_center_y}, cesium_center_z={s.cesium_center_z}")
    print()
    
    # 2. 查看设备信息
    print("【2. 监测设备信息】")
    devices = MonitoringDevice.objects.all()
    print(f"  设备总数: {devices.count()}")
    
    # 按设备类型分组统计
    device_types = {
        'inverted_plumb_left_right': '倒垂线-左右岸位移',
        'inverted_plumb_up_down': '倒垂线-上下游位移',
        'hydrostatic_leveling': '静力水准仪',
        'tension_wire_up_down': '张引线',
        'water_level_upstream': '水位检测器-上游',
        'water_level_downstream': '水位检测器-下游',
    }
    
    for dtype, dtype_name in device_types.items():
        count = devices.filter(device_type=dtype).count()
        if count > 0:
            print(f"  - {dtype_name}: {count} 个")
            # 显示前3个设备示例
            samples = devices.filter(device_type=dtype)[:3]
            for dev in samples:
                install_time = dev.install_time.strftime("%Y-%m-%d") if dev.install_time else "未设置"
                print(f"    · {dev.device_name} (状态: {dev.device_status}, 安装时间: {install_time})")
    print()
    
    # 3. 查看测点信息
    print("【3. 监测测点信息】")
    points = Point.objects.all()
    print(f"  测点总数: {points.count()}")
    # 显示前10个测点示例
    sample_points = points[:10]
    for p in sample_points:
        print(f"  - {p.point_code} (所属设备: {p.device.device_name}, 类型: {p.device.get_device_type_display()})")
    if points.count() > 10:
        print(f"  ... 还有 {points.count() - 10} 个测点")
    print()
    
    # 4. 查看监测数据信息
    print("【4. 监测数据信息】")
    data = MonitorData.objects.all()
    print(f"  监测数据总数: {data.count()}")
    
    # 按测量类型统计
    print("  按测量类型统计:")
    plumb_lr_count = data.filter(inverted_plumb_left_right__isnull=False).count()
    plumb_ud_count = data.filter(inverted_plumb_up_down__isnull=False).count()
    tension_count = data.filter(tension_wire_up_down__isnull=False).count()
    leveling_count = data.filter(hydrostatic_leveling_settlement__isnull=False).count()
    water_up_count = data.filter(water_level_upstream__isnull=False).count()
    water_down_count = data.filter(water_level_downstream__isnull=False).count()
    
    print(f"    - 倒垂线左右岸位移: {plumb_lr_count} 条")
    print(f"    - 倒垂线上下游位移: {plumb_ud_count} 条")
    print(f"    - 张引线上下游位移: {tension_count} 条")
    print(f"    - 静力水准沉降: {leveling_count} 条")
    print(f"    - 上游水位: {water_up_count} 条")
    print(f"    - 下游水位: {water_down_count} 条")
    
    # 统计异常数据 (-999.x)
    print("\n  异常数据标记统计:")
    anomaly_999_1 = data.filter(
        Q(inverted_plumb_left_right=-999.1) | Q(inverted_plumb_up_down=-999.1) |
        Q(tension_wire_up_down=-999.1) | Q(hydrostatic_leveling_settlement=-999.1) |
        Q(water_level_upstream=-999.1) | Q(water_level_downstream=-999.1)
    ).count()
    anomaly_999_2 = data.filter(
        Q(inverted_plumb_left_right=-999.2) | Q(inverted_plumb_up_down=-999.2) |
        Q(tension_wire_up_down=-999.2) | Q(hydrostatic_leveling_settlement=-999.2) |
        Q(water_level_upstream=-999.2) | Q(water_level_downstream=-999.2)
    ).count()
    anomaly_999_9 = data.filter(
        Q(inverted_plumb_left_right=-999.9) | Q(inverted_plumb_up_down=-999.9) |
        Q(tension_wire_up_down=-999.9) | Q(hydrostatic_leveling_settlement=-999.9) |
        Q(water_level_upstream=-999.9) | Q(water_level_downstream=-999.9)
    ).count()
    
    print(f"    - -999.1 (低于标尺水位): {anomaly_999_1} 条")
    print(f"    - -999.2 (被遮挡无法观测): {anomaly_999_2} 条")
    print(f"    - -999.9 (乱码数据): {anomaly_999_9} 条")
    
    if anomaly_999_1 + anomaly_999_2 + anomaly_999_9 > 0:
        print(f"\n  ⚠️  共有 {anomaly_999_1 + anomaly_999_2 + anomaly_999_9} 条数据需要后期手动修正")
    
    # 显示最新5条数据
    print("\n  最新5条监测数据:")
    latest_data = data.order_by('-monitor_time')[:5]
    for d in latest_data:
        time_str = d.monitor_time.strftime("%Y-%m-%d %H:%M:%S")
        fields = []
        if d.inverted_plumb_left_right is not None:
            fields.append(f"左右={d.inverted_plumb_left_right}")
        if d.inverted_plumb_up_down is not None:
            fields.append(f"上下={d.inverted_plumb_up_down}")
        if d.hydrostatic_leveling_settlement is not None:
            fields.append(f"水准={d.hydrostatic_leveling_settlement}")
        if d.tension_wire_up_down is not None:
            fields.append(f"引张={d.tension_wire_up_down}")
        if d.water_level_upstream is not None:
            fields.append(f"上游水位={d.water_level_upstream}")
        if d.water_level_downstream is not None:
            fields.append(f"下游水位={d.water_level_downstream}")
        print(f"    - {time_str} | {d.point.point_code} | {', '.join(fields)}")
    
    print("\n" + "="*60)
    print("✅ 数据统计完成！")
    print("="*60 + "\n")


def step0_clear_data():
    """
    清空步骤：删除所有监测数据，但保留Structure/MonitoringDevice/Point
    用于重新导入时的数据清理
    """
    print("\n===== 清除监测数据 =====")
    try:
        count = MonitorData.objects.all().count()
        MonitorData.objects.all().delete()
        print(f"✅ 已清除 {count} 条监测数据")
    except Exception as e:
        print(f"❌ 清除数据失败：{e}")


if __name__ == "__main__":
    # Excel 文件路径（脚本在 backend 目录，Excel 在上级目录的课设材料文件夹）
    EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "课设材料", "监测资料.xlsx")
    
    print("=" * 80)
    print("智慧水利监测数据导入工具")
    print("=" * 80)
    print("\n请选择要执行的步骤：")
    print("  0 - 清除所有监测数据 (step0)")
    print("  1 - 导入大坝基础信息 (step1)")
    print("  2 - 导入监测设备 (step2)")
    print("  3 - 导入测点 (step3)")
    print("  4 - 导入监测数据 (step4)")
    print("  5 - 查看导入数据统计 (step5)")
    print("  all - 执行全部步骤")
    print("  clean - 清除数据后执行步骤4和5")
    print("  quit - 退出")
    print()
    
    # 支持命令行参数选择步骤：python import_monitor_data.py 4/5/all/clean
    if len(sys.argv) > 1:
        choice = sys.argv[1].strip().lower()
        print(f"(命令行选择) 执行步骤: {choice}")
    else:
        choice = input("请输入选项 (0/1/2/3/4/5/all/clean/quit): ").strip().lower()
    
    if choice == "quit":
        print("已退出")
        exit(0)
    
    # Step 0: 清除数据
    if choice in ["0", "clean"]:
        step0_clear_data()
        if choice == "0":
            exit(0)
    
    # Step 1: 导入大坝
    if choice in ["1", "all"]:
        step1_import_structure(EXCEL_PATH)
        if choice == "1":
            exit(0)
    
    # Step 2: 导入设备
    if choice in ["2", "all"]:
        step2_import_devices(EXCEL_PATH)
        if choice == "2":
            exit(0)
    
    # Step 3: 导入测点
    if choice in ["3", "all"]:
        step3_import_points(EXCEL_PATH)
        if choice == "3":
            exit(0)
    
    # Step 4: 导入监测数据
    if choice in ["4", "all", "clean"]:
        # 为每个 sheet 定义布局配置
        sheets_config = [
            # 倒垂线（双行标题）
            ("倒垂线", "inverted_plumb_left_right", 2, 6, 7, 4, 5),
            ("倒垂线", "inverted_plumb_up_down",    2, 6, 7, 4, 5),
            
            # 静力水准
            ("静力水准", "hydrostatic_leveling_settlement", 2, 7, 8, 3, 6),
            
            # 张引线
            ("引张线", "tension_wire_up_down", 1, 7, 8, 4, 6),
            
            # 水位
            ("水位", "water_level_upstream",   1, 2, 3, 2),
            ("水位", "water_level_downstream", 1, 2, 3, 3),
        ]
        
        for config in sheets_config:
            if len(config) == 7:
                # 有 device_name_row 参数（双行标题）
                sheet_name, device_type_field, time_col, header_row, data_start_row, data_col_start, device_name_row = config
                step4_import_monitor_data(EXCEL_PATH, sheet_name, device_type_field, time_col, header_row, data_start_row, data_col_start, device_name_row)
            else:
                # 没有 device_name_row 参数（单行标题）
                sheet_name, device_type_field, time_col, header_row, data_start_row, data_col_start = config
                step4_import_monitor_data(EXCEL_PATH, sheet_name, device_type_field, time_col, header_row, data_start_row, data_col_start)
        
        print("\n" + "=" * 80)
        print("所有数据导入完成！")
        print("=" * 80)
        if choice == "4":
            exit(0)
    
    # Step 5: 查看数据统计
    if choice in ["5", "clean"]:
        step5_view_data()
        exit(0)